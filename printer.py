from termcolor import cprint
import openpyxl as xl
import random
from matrix_maps import get_data, get_times, get_distances
from urllib.parse import quote


LOCATION = ", Región Metropolitana, Chile"

wb = xl.load_workbook(filename = 'recorridos.xlsx')
ws1 = wb['Tiempo comunas']
ws2 = wb['Distancia comunas']

#print(ws['Z9'].fill.start_color.index)

places = ""

for col in range(2, 44):
    place = ws1.cell(column=col, row=1).value + LOCATION
    places += place + "|"
places = places[:-1]


data = get_data(quote(places), quote(places))


cprint("[STATUS] " + data['status'], "blue")

times = get_times(data)
distances = get_distances(data)

for row in range(2, 44):
    for col in range(2, 44):
        ws1.cell(column=col, row=row, value=times[row][col])
        ws2.cell(column=col, row=row, value=distances[row][col])

wb.save(filename = 'example.xlsx')

print()
print(times)
print()
print(distances)
print()
